import re
from io import BytesIO

from openpyxl import load_workbook

from app.domain.entities import Voter
from app.domain.interfaces import IExcelImporter

REQUIRED_COLUMNS = {"email", "telephone", "nom", "prenom"}
PHONE_PATTERN = re.compile(r"^\+7\d{10}$")


def _digits_from_cell(raw: object) -> str:
    text = str(raw).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\D", "", text)


def _normalize_phone(raw: object) -> str:
    digits = _digits_from_cell(raw)
    if not digits:
        raise ValueError(f"Numéro russe invalide : {raw}")

    # 8XXXXXXXXXX (11 chiffres) → 7XXXXXXXXXX
    if digits.startswith("8") and len(digits) == 11:
        digits = "7" + digits[1:]

    # 8XXXXXXXXX (10 chiffres) → 7XXXXXXXXX
    if digits.startswith("8") and len(digits) == 10:
        digits = "7" + digits[1:]

    # XXXXXXXXXX (10 chiffres sans indicatif) → 7XXXXXXXXXX
    if len(digits) == 10 and not digits.startswith("7"):
        digits = "7" + digits

    # 89XXXXXXX / 79XXXXXXX (9 chiffres — fréquent quand Excel tronque)
    if len(digits) == 9:
        if digits.startswith("8"):
            digits = "7" + digits[1:]
        elif digits.startswith("9"):
            digits = "7" + digits

    # Compléter jusqu'à 11 chiffres (7 + 10 chiffres nationaux)
    if digits.startswith("7") and len(digits) < 11:
        digits = digits.ljust(11, "0")
    elif digits.startswith("7") and len(digits) > 11:
        raise ValueError(f"Numéro russe invalide : {raw}")

    if digits.startswith("7") and len(digits) == 11:
        return f"+{digits}"

    raise ValueError(f"Numéro russe invalide : {raw}")


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace("é", "e").replace(" ", "_")


class ExcelVoterImporter(IExcelImporter):
    async def parse_voters(self, file_content: bytes) -> list[Voter]:
        workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Fichier Excel vide")

        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            raise ValueError("Fichier Excel sans en-têtes")

        column_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                column_map[_normalize_header(str(cell))] = idx

        missing = REQUIRED_COLUMNS - set(column_map.keys())
        if missing:
            raise ValueError(f"Colonnes manquantes : {', '.join(sorted(missing))}")

        voters: list[Voter] = []
        errors: list[str] = []

        for row_num, row in enumerate(rows, start=2):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            try:
                email = str(row[column_map["email"]]).strip().lower()
                phone = _normalize_phone(str(row[column_map["telephone"]]))
                last_name = str(row[column_map["nom"]]).strip()
                first_name = str(row[column_map["prenom"]]).strip()

                if not email or "@" not in email:
                    raise ValueError("email invalide")
                if not PHONE_PATTERN.match(phone):
                    raise ValueError(f"téléphone invalide après normalisation : {phone}")
                if not last_name or not first_name:
                    raise ValueError("nom ou prénom vide")

                voters.append(
                    Voter(
                        id=None,
                        email=email,
                        phone=phone,
                        last_name=last_name,
                        first_name=first_name,
                    )
                )
            except (ValueError, TypeError, IndexError) as exc:
                errors.append(f"Ligne {row_num} : {exc}")

        workbook.close()

        if errors:
            raise ValueError("\n".join(errors[:20]))

        if not voters:
            raise ValueError("Aucun électeur valide trouvé dans le fichier")

        return voters
